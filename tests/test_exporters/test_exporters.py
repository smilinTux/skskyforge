"""
Tests for PDF, CSV, and Excel exporters.

Copyright (C) 2025 S&K Holding QT (Quantum Technologies)
SK = staycuriousANDkeepsmilin
"""

import csv
import pytest
from datetime import date, timedelta
from pathlib import Path

from skskyforge.generators import generate_daily_entry
from skskyforge.exporters import export_csv, export_excel, export_pdf
from skskyforge.models import BirthData, UserProfile


@pytest.fixture
def sample_profile():
    return UserProfile(
        name="export-test",
        birth_data=BirthData(date=date(1990, 6, 15)),
    )


@pytest.fixture
def sample_entries(sample_profile):
    """Generate three days of entries."""
    start = date(2026, 3, 1)
    return [
        generate_daily_entry(start + timedelta(days=i), sample_profile)
        for i in range(3)
    ]


class TestCsvExporter:

    def test_creates_file(self, sample_entries, tmp_path):
        out = tmp_path / "test.csv"
        result = export_csv(sample_entries, out)
        assert result.exists()
        assert result.stat().st_size > 0

    def test_correct_row_count(self, sample_entries, tmp_path):
        out = tmp_path / "test.csv"
        export_csv(sample_entries, out)
        with open(out) as f:
            reader = csv.reader(f)
            rows = list(reader)
        # header + 3 data rows
        assert len(rows) == 4

    def test_header_contains_core_fields(self, sample_entries, tmp_path):
        out = tmp_path / "test.csv"
        export_csv(sample_entries, out)
        with open(out) as f:
            reader = csv.DictReader(f)
            fields = reader.fieldnames
        for field in ("date", "day_of_week", "daily_theme", "moon_phase", "moon_sign"):
            assert field in fields

    def test_dates_in_order(self, sample_entries, tmp_path):
        out = tmp_path / "test.csv"
        export_csv(sample_entries, out)
        with open(out) as f:
            reader = csv.DictReader(f)
            dates = [row["date"] for row in reader]
        assert dates == sorted(dates)


class TestExcelExporter:

    def test_creates_file(self, sample_entries, tmp_path):
        out = tmp_path / "test.xlsx"
        result = export_excel(sample_entries, out)
        assert result.exists()
        assert result.stat().st_size > 0

    def test_has_overview_sheet(self, sample_entries, tmp_path):
        from openpyxl import load_workbook
        out = tmp_path / "test.xlsx"
        export_excel(sample_entries, out)
        wb = load_workbook(out)
        assert "Overview" in wb.sheetnames

    def test_overview_row_count(self, sample_entries, tmp_path):
        from openpyxl import load_workbook
        out = tmp_path / "test.xlsx"
        export_excel(sample_entries, out)
        wb = load_workbook(out)
        ws = wb["Overview"]
        # header + 3 data rows
        assert ws.max_row == 4


class TestPdfExporter:

    def test_creates_file(self, sample_entries, tmp_path):
        out = tmp_path / "test.pdf"
        result = export_pdf(sample_entries, out)
        assert result.exists()
        assert result.stat().st_size > 0

    def test_file_starts_with_pdf_magic(self, sample_entries, tmp_path):
        out = tmp_path / "test.pdf"
        export_pdf(sample_entries, out)
        with open(out, "rb") as f:
            magic = f.read(5)
        assert magic == b"%PDF-"

    def test_single_entry_works(self, sample_entries, tmp_path):
        out = tmp_path / "single.pdf"
        export_pdf(sample_entries[:1], out)
        assert out.exists()
