"""
Excel exporter for SKSkyforge daily entries.

Produces an .xlsx workbook with styled sheets for overview,
moon phases, numerology, biorhythm, and planetary data.

Copyright (C) 2025 S&K Holding QT (Quantum Technologies)
Licensed under AGPL-3.0.
"""

from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..models.daily_entry import DailyPreparation


# Sovereign theme colours
_HEADER_FILL = PatternFill(start_color="1A1A35", end_color="1A1A35", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", bold=True, color="A78BFA", size=11)
_DATE_FONT = Font(name="Calibri", bold=True, color="00E5FF", size=11)
_BODY_FONT = Font(name="Calibri", color="E2E8F0", size=10)


def _apply_header(ws, headers: list[str]) -> None:
    """Write styled header row."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _auto_width(ws) -> None:
    """Auto-size columns based on content."""
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            val = str(row[0]) if row[0] is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[letter].width = min(max_len + 3, 40)


def _write_overview(wb: Workbook, entries: List[DailyPreparation]) -> None:
    """Write the Overview sheet with daily summaries."""
    ws = wb.active
    ws.title = "Overview"

    headers = [
        "Date", "Day", "Theme", "Moon Phase", "Moon Sign",
        "Sun Sign", "Personal Day", "Energy", "Risk", "Affirmation",
    ]
    _apply_header(ws, headers)

    for row_idx, entry in enumerate(entries, 2):
        ws.cell(row=row_idx, column=1, value=entry.date.isoformat()).font = _DATE_FONT
        ws.cell(row=row_idx, column=2, value=entry.day_of_week)
        ws.cell(row=row_idx, column=3, value=entry.daily_theme)
        ws.cell(row=row_idx, column=4, value=entry.moon.phase)
        ws.cell(row=row_idx, column=5, value=entry.moon.zodiac_sign)
        ws.cell(row=row_idx, column=6, value=entry.solar_transit.sun_sign)
        ws.cell(row=row_idx, column=7, value=entry.numerology.personal_day)
        ws.cell(row=row_idx, column=8, value=entry.biorhythm.overall_energy)
        ws.cell(row=row_idx, column=9, value=entry.risk_analysis.overall_risk_level)
        ws.cell(row=row_idx, column=10, value=entry.affirmation)

    _auto_width(ws)


def _write_moon(wb: Workbook, entries: List[DailyPreparation]) -> None:
    """Write Moon sheet with phase and sign details."""
    ws = wb.create_sheet("Moon")

    headers = [
        "Date", "Phase", "Illumination %", "Sign", "Element",
        "Modality", "VOC", "Energy Theme",
    ]
    _apply_header(ws, headers)

    for row_idx, entry in enumerate(entries, 2):
        m = entry.moon
        ws.cell(row=row_idx, column=1, value=entry.date.isoformat()).font = _DATE_FONT
        ws.cell(row=row_idx, column=2, value=m.phase)
        ws.cell(row=row_idx, column=3, value=m.phase_percentage)
        ws.cell(row=row_idx, column=4, value=m.zodiac_sign)
        ws.cell(row=row_idx, column=5, value=m.sign_element)
        ws.cell(row=row_idx, column=6, value=m.sign_modality)
        ws.cell(row=row_idx, column=7, value="Yes" if m.moon_void_of_course else "")
        ws.cell(row=row_idx, column=8, value=m.energy_theme)

    _auto_width(ws)


def _write_numerology(wb: Workbook, entries: List[DailyPreparation]) -> None:
    """Write Numerology sheet."""
    ws = wb.create_sheet("Numerology")

    headers = [
        "Date", "Life Path", "Personal Year", "Personal Month",
        "Personal Day", "Universal Day", "Theme", "Energy Quality",
    ]
    _apply_header(ws, headers)

    for row_idx, entry in enumerate(entries, 2):
        n = entry.numerology
        ws.cell(row=row_idx, column=1, value=entry.date.isoformat()).font = _DATE_FONT
        ws.cell(row=row_idx, column=2, value=n.life_path)
        ws.cell(row=row_idx, column=3, value=n.personal_year)
        ws.cell(row=row_idx, column=4, value=n.personal_month)
        ws.cell(row=row_idx, column=5, value=n.personal_day)
        ws.cell(row=row_idx, column=6, value=n.universal_day)
        ws.cell(row=row_idx, column=7, value=n.day_theme)
        ws.cell(row=row_idx, column=8, value=n.energy_quality)

    _auto_width(ws)


def _write_biorhythm(wb: Workbook, entries: List[DailyPreparation]) -> None:
    """Write Biorhythm sheet."""
    ws = wb.create_sheet("Biorhythm")

    headers = [
        "Date", "Physical", "Emotional", "Intellectual", "Overall Energy",
    ]
    _apply_header(ws, headers)

    for row_idx, entry in enumerate(entries, 2):
        b = entry.biorhythm
        ws.cell(row=row_idx, column=1, value=entry.date.isoformat()).font = _DATE_FONT
        ws.cell(row=row_idx, column=2, value=round(b.physical, 1))
        ws.cell(row=row_idx, column=3, value=round(b.emotional, 1))
        ws.cell(row=row_idx, column=4, value=round(b.intellectual, 1))
        ws.cell(row=row_idx, column=5, value=b.overall_energy)

    _auto_width(ws)


def export_excel(
    entries: List[DailyPreparation],
    output_path: Path,
) -> Path:
    """
    Export daily entries to Excel (.xlsx) format.

    Creates a workbook with four sheets: Overview, Moon, Numerology,
    and Biorhythm.

    Args:
        entries: List of generated daily entries.
        output_path: Path for the output .xlsx file.

    Returns:
        Path to the written file.
    """
    wb = Workbook()

    _write_overview(wb, entries)
    _write_moon(wb, entries)
    _write_numerology(wb, entries)
    _write_biorhythm(wb, entries)

    wb.save(str(output_path))
    return output_path
